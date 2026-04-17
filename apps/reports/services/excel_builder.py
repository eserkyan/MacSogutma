from __future__ import annotations

from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.core.services.status_labels import get_test_status_label
from apps.reports.services.report_context import ReportContextService
from apps.tests.models import TestRecord


class ExcelBuilderService:
    def build(self, test_record: TestRecord, language: str = "tr") -> str:
        target = self.target_path_for(test_record, language=language)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(self.build_bytes(test_record, language=language))
        test_record.excel_file_path = str(target)
        test_record.excel_generated_at = timezone.now()
        test_record.save(update_fields=["excel_file_path", "excel_generated_at", "updated_at"])
        return str(target)

    def target_path_for(self, test_record: TestRecord, language: str = "tr") -> Path:
        stamp = timezone.localtime()
        root = Path(settings.PLC_CONFIG["report_root_path"])
        return (
            root
            / test_record.company.code
            / test_record.product_model.model_code
            / f"{stamp:%Y}"
            / f"{stamp:%m}"
            / f"{test_record.test_no}_{language}.xlsx"
        )

    def build_bytes(self, test_record: TestRecord, language: str = "tr") -> bytes:
        context = ReportContextService().build(test_record, language=language)
        workbook = Workbook()
        self._build_summary_sheet(workbook.active, context)
        self._build_evaluation_sheet(workbook.create_sheet("Evaluation"), context)
        phase_sections = {str(section.get("slug")): section for section in context.get("phase_stat_sections", [])}
        self._build_phase_stat_sheet(
            workbook.create_sheet("Start Stats"),
            phase_sections.get("start", {}).get("rows", []),
        )
        stable_rows = [
            {
                "parameter_name": row["parameter_name"],
                "parameter_code": row["parameter_code"],
                "unit": row["unit"],
                "stats": {
                    "min_value": row["measured_min"],
                    "avg_value": row["avg_value"],
                    "max_value": row["measured_max"],
                },
            }
            for row in context.get("evaluation_rows", [])
        ]
        self._build_phase_stat_sheet(workbook.create_sheet("Stable Stats"), stable_rows)
        self._build_phase_stat_sheet(
            workbook.create_sheet("Stop Stats"),
            phase_sections.get("stop", {}).get("rows", []),
        )
        self._build_limit_sheet(workbook.create_sheet("Limit Excursions"), context)
        self._build_sample_sheet(workbook.create_sheet("All Samples"), context["export_sample_columns"], context["all_samples"])
        self._build_sample_sheet(workbook.create_sheet("Start Samples"), context["export_sample_columns"], context["samples_by_phase"]["start"])
        self._build_sample_sheet(workbook.create_sheet("Stable Samples"), context["export_sample_columns"], context["samples_by_phase"]["stable"])
        self._build_sample_sheet(workbook.create_sheet("Stop Samples"), context["export_sample_columns"], context["samples_by_phase"]["stop"])

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _build_summary_sheet(self, sheet, context: dict[str, object]) -> None:
        test_record = context["test_record"]
        sheet.title = "Summary"
        language = context.get("language", "tr")
        rows = [
            ("Test No", test_record.test_no),
            ("Status", get_test_status_label(test_record.status, context.get("language", "tr"))),
            ("Company", test_record.company.name),
            ("Model", test_record.product_model.model_name),
            ("Recipe", test_record.recipe_name_snapshot),
            ("Start Duration", test_record.start_duration_sec_snapshot),
            ("Stable Duration", test_record.stable_duration_sec_snapshot),
            ("Stop Duration", test_record.stop_duration_sec_snapshot),
            ("PDF Path", str(test_record.pdf_file_path or "-")),
            ("Excel Path", str(test_record.excel_file_path or self.target_path_for(test_record, language=language))),
        ]
        for index, (label, value) in enumerate(rows, start=1):
            sheet.cell(index, 1, label).font = Font(bold=True)
            sheet.cell(index, 2, value)
        status_value = str(test_record.status)
        status_cell = sheet.cell(2, 2)
        if status_value == "COMPLETED_PASS":
            status_cell.fill = PatternFill("solid", fgColor="16A34A")
            status_cell.font = Font(bold=True, color="FFFFFF", size=12)
        elif status_value in {"COMPLETED_FAIL", "FAILED_TO_START", "ABORTED"}:
            status_cell.fill = PatternFill("solid", fgColor="DC2626")
            status_cell.font = Font(bold=True, color="FFFFFF", size=12)
        else:
            status_cell.fill = PatternFill("solid", fgColor="64748B")
            status_cell.font = Font(bold=True, color="FFFFFF", size=12)
        status_cell.alignment = Alignment(horizontal="center")
        sheet.cell(len(rows) + 2, 1, "Traceability Note").font = Font(bold=True)
        sheet.cell(len(rows) + 3, 1, "All phases are included for technical traceability.")
        sheet.cell(len(rows) + 4, 1, "Acceptance criteria are evaluated using phase average values for phases with active limits.")

    def _build_evaluation_sheet(self, sheet, context: dict[str, object]) -> None:
        headers = [
            "Parameter",
            "Code",
            "Unit",
            "Measured Min",
            "AVG",
            "Measured Max",
            "Limit Min",
            "Limit Max",
            "Status",
            "Message",
        ]
        self._write_header(sheet, headers)
        for row_index, row in enumerate(context["evaluation_rows"], start=2):
            sheet.cell(row_index, 1, row["parameter_name"])
            sheet.cell(row_index, 2, row["parameter_code"])
            sheet.cell(row_index, 3, row["unit"])
            sheet.cell(row_index, 4, self._num(row["measured_min"]))
            sheet.cell(row_index, 5, self._num(row["avg_value"]))
            sheet.cell(row_index, 6, self._num(row["measured_max"]))
            sheet.cell(row_index, 7, self._num(row["limit_min"]))
            sheet.cell(row_index, 8, self._num(row["limit_max"]))
            passed = row["passed"]
            status_cell = sheet.cell(row_index, 9, "PASS" if passed else "FAIL")
            status_cell.fill = PatternFill("solid", fgColor="16A34A" if passed else "DC2626")
            status_cell.font = Font(bold=True, color="FFFFFF")
            status_cell.alignment = Alignment(horizontal="center")
            sheet.cell(row_index, 10, row["message"])

    def _build_limit_sheet(self, sheet, context: dict[str, object]) -> None:
        headers = ["Parameter", "Code", "Phase", "Elapsed Sec", "Measured", "Unit", "Limit Type", "Limit", "Message"]
        self._write_header(sheet, headers)
        for row_index, row in enumerate(context["limit_rows"], start=2):
            sheet.cell(row_index, 1, row["parameter_name"])
            sheet.cell(row_index, 2, row["parameter_code"])
            sheet.cell(row_index, 3, row["phase_name"])
            sheet.cell(row_index, 4, self._num(row["elapsed_seconds"]))
            sheet.cell(row_index, 5, self._num(row["sample_value"]))
            sheet.cell(row_index, 6, row["unit"])
            sheet.cell(row_index, 7, row["limit_type"])
            sheet.cell(row_index, 8, self._num(row["limit_value"]))
            sheet.cell(row_index, 9, row["message"])

    def _build_phase_stat_sheet(self, sheet, rows: list[dict[str, object]]) -> None:
        headers = ["Parameter", "Code", "Unit", "Measured Min", "AVG", "Measured Max", "Limit Min", "Limit Max", "Status", "Message"]
        self._write_header(sheet, headers)
        for row_index, row in enumerate(rows, start=2):
            stats = row["stats"]
            sheet.cell(row_index, 1, row["parameter_name"])
            sheet.cell(row_index, 2, row["parameter_code"])
            sheet.cell(row_index, 3, row["unit"])
            sheet.cell(row_index, 4, self._num(stats["min_value"]))
            sheet.cell(row_index, 5, self._num(stats["avg_value"]))
            sheet.cell(row_index, 6, self._num(stats["max_value"]))
            sheet.cell(row_index, 7, self._num(row.get("min_limit")) if row.get("min_enabled") else None)
            sheet.cell(row_index, 8, self._num(row.get("max_limit")) if row.get("max_enabled") else None)
            passed = row.get("passed")
            status = "N/A" if passed is None else "PASS" if passed else "FAIL"
            status_cell = sheet.cell(row_index, 9, status)
            if passed is not None:
                status_cell.fill = PatternFill("solid", fgColor="16A34A" if passed else "DC2626")
                status_cell.font = Font(bold=True, color="FFFFFF")
            else:
                status_cell.fill = PatternFill("solid", fgColor="CBD5E1")
                status_cell.font = Font(bold=True, color="334155")
            status_cell.alignment = Alignment(horizontal="center")
            sheet.cell(row_index, 10, row.get("display_message") or row.get("message") or "")

    def _build_sample_sheet(self, sheet, columns: list[tuple[str, str]], rows: list[dict[str, object]]) -> None:
        self._write_header(sheet, [label for _, label in columns])
        for row_index, row in enumerate(rows, start=2):
            for col_index, (field_name, _) in enumerate(columns, start=1):
                sheet.cell(row_index, col_index, self._cell_value(row.get(field_name)))

    @staticmethod
    def _write_header(sheet, headers: list[str]) -> None:
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_index, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
            sheet.column_dimensions[cell.column_letter].width = max(14, min(24, len(header) + 2))

    @staticmethod
    def _num(value: object) -> float | None:
        if value is None or value == "":
            return None
        return round(float(value), 2)

    def _cell_value(self, value: object) -> object:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return round(float(value), 2)
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return value
